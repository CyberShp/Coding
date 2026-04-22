"""
Array management API — CRUD, batch, import, connect/disconnect, metrics.

Sub-modules included here (all endpoints still served under /arrays):
  array_status.py    — status cache, active issues, /status /statuses /search /watchers
  array_alert_sync.py — sync logic, /refresh
  array_agent_ops.py  — agent deploy/start/stop/restart/logs/agent-config

Backward-compat re-exports are at the bottom so existing importers
(main.py, core/alert_sync.py, acknowledgements.py, etc.) keep working.
"""

import asyncio
import json
import logging
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, status, Body, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import get_config
from ..core.agent_deployer import AgentDeployer
from ..core.ssh_pool import get_ssh_pool, SSHPool
from ..core.system_alert import sys_error, sys_info
from ..db.database import get_db, AsyncSessionLocal
from ..models.array import (
    ArrayModel, ArrayCreate, ArrayUpdate, ArrayResponse,
    ArrayStatus, ConnectionState,
)

from .array_status import (
    _array_status_cache,
    _get_array_status,
    _get_array_or_404,
    _resolve_ips_to_nicknames,
    _expand_l1_tag_filter,
    status_router,
)
from .array_alert_sync import sync_array_alerts, sync_router
from .array_agent_ops import agent_router

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/arrays", tags=["arrays"])

# Include sub-routers (all relative paths, prefix="/arrays" comes from this router)
router.include_router(status_router)
router.include_router(sync_router)
router.include_router(agent_router)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

async def _run_blocking(func, _timeout: float, *args, **kwargs):
    """Run sync I/O in threadpool to avoid blocking event loop.

    _timeout is deliberately underscore-prefixed so that **kwargs can forward
    a ``timeout`` keyword to the wrapped *func* without colliding.
    """
    loop = asyncio.get_running_loop()
    return await asyncio.wait_for(
        loop.run_in_executor(None, lambda: func(*args, **kwargs)),
        timeout=_timeout,
    )


class BatchActionRequest(BaseModel):
    """Request model for batch operations"""
    array_ids: List[str]
    password: Optional[str] = None  # For batch connect


# ---------------------------------------------------------------------------
# Import helpers
# ---------------------------------------------------------------------------

# High-contrast color pool for auto-created tags during import
_IMPORT_TAG_COLORS = [
    "#409EFF", "#67C23A", "#E6A23C", "#F56C6C", "#909399",
    "#00BCD4", "#9C27B0", "#FF5722", "#795548", "#607D8B",
    "#E91E63", "#3F51B5", "#009688", "#FF9800", "#8BC34A",
    "#CDDC39", "#03A9F4", "#673AB7", "#F44336", "#4CAF50",
]


def _parse_import_file(content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Parse CSV or Excel file into list of dicts."""
    import csv
    import io

    name_lower = (filename or "").lower()
    rows = []

    if name_lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append({k.strip() if k else k: (v.strip() if v else "") for k, v in (r or {}).items()})
    elif name_lower.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Excel support requires openpyxl. Install with: pip install openpyxl",
            )
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no sheets")
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2):
            vals = [str(c.value or "").strip() if c.value is not None else "" for c in row]
            rows.append(dict(zip(headers, vals)))
        wb.close()
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported format. Use .csv or .xlsx",
        )
    return rows


def _normalize_import_row(row: Dict[str, Any]) -> tuple[Optional[Dict[str, Any]], str]:
    """Normalize row to array fields. Returns (None, reason) if row is invalid."""
    def get(key_variants, default=""):
        for k in key_variants:
            for rk, rv in row.items():
                if (rk or "").strip().lower() == k.lower():
                    return (rv or "").strip()
        return default

    name = get(["name", "名称", "array_name"])
    host = get(["host", "ip", "地址", "hostname"])
    if not name and not host:
        return None, "name_and_host_empty"
    if not name:
        return None, "name_empty"
    if not host:
        return None, "host_empty"
    port_s = get(["port", "端口"], "22")
    try:
        port = int(port_s) if port_s else 22
    except ValueError:
        port = 22
    username = get(["username", "user", "用户名"], "root")
    password = get(["password", "密码"], "")
    tag_name = get(["tag", "标签", "tag_name"])
    tag_l1 = get(["tag_l1", "一级标签", "tag_l1_name"])
    tag_l2 = get(["tag_l2", "二级标签", "tag_l2_name"])
    color = get(["color", "颜色"])
    if color and (not color.startswith("#") or len(color) not in (4, 7)):
        color = ""
    return {
        "name": name,
        "host": host,
        "port": port,
        "username": username,
        "password": password,
        "tag_name": tag_name or None,
        "tag_l1": tag_l1 or None,
        "tag_l2": tag_l2 or None,
        "color": color or None,
    }, ""


# ---------------------------------------------------------------------------
# CRUD endpoints
# ---------------------------------------------------------------------------

@router.get("", response_model=List[ArrayResponse])
async def list_arrays(
    tag_id: Optional[int] = Query(None, description="Filter by tag ID"),
    db: AsyncSession = Depends(get_db),
):
    """Get all arrays (database records only, no connection state)"""
    from ..models.tag import TagModel

    query = select(ArrayModel)
    if tag_id is not None:
        query = await _expand_l1_tag_filter(query, tag_id, db)

    result = await db.execute(query)
    arrays = result.scalars().all()

    tag_ids = {a.tag_id for a in arrays if a.tag_id}
    tags_map = {}
    parent_map = {}
    if tag_ids:
        tag_result = await db.execute(select(TagModel).where(TagModel.id.in_(tag_ids)))
        tags_map = {t.id: t for t in tag_result.scalars().all()}
        parent_ids = {t.parent_id for t in tags_map.values() if t.parent_id}
        if parent_ids:
            parent_result = await db.execute(
                select(TagModel.id, TagModel.name).where(TagModel.id.in_(parent_ids))
            )
            parent_map = {r[0]: r[1] for r in parent_result.all()}

    def _l1_l2(tag):
        if not tag:
            return None, None
        if tag.level == 2 and tag.parent_id:
            return parent_map.get(tag.parent_id), tag.name
        if tag.level == 1:
            return tag.name, None
        return None, tag.name

    responses = []
    for arr in arrays:
        tag = tags_map.get(arr.tag_id) if arr.tag_id else None
        l1, l2 = _l1_l2(tag)
        responses.append(ArrayResponse(
            id=arr.id,
            array_id=arr.array_id,
            name=arr.name,
            host=arr.host,
            port=arr.port,
            username=arr.username,
            key_path=arr.key_path or "",
            folder=arr.folder or "",
            tag_id=arr.tag_id,
            tag_name=tag.name if tag else None,
            tag_color=tag.color if tag else None,
            tag_l1_name=l1,
            tag_l2_name=l2,
            created_at=arr.created_at,
            updated_at=arr.updated_at,
        ))

    return responses


@router.post("/import")
async def import_arrays(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    ssh_pool: SSHPool = Depends(get_ssh_pool),
):
    """Import arrays from CSV or Excel. Columns: name, host, port?, username?, tag?, tag_l1?, tag_l2?, color?"""
    from ..models.tag import TagModel
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File too large (max 10MB)")
    filename = file.filename or ""
    rows = _parse_import_file(content, filename)
    tag_cache: Dict[tuple, int] = {}
    color_idx = [0]

    async def get_or_create_tag_from_import(norm: Dict[str, Any]) -> Optional[int]:
        tag_l1 = (norm.get("tag_l1") or "").strip()
        tag_l2 = (norm.get("tag_l2") or norm.get("tag_name") or "").strip()
        if not tag_l2 and tag_l1:
            tag_l2 = tag_l1
            tag_l1 = ""
        if not tag_l2:
            return None
        color = (norm.get("color") or "").strip()
        if not color or not color.startswith("#") or len(color) not in (4, 7):
            color = _IMPORT_TAG_COLORS[color_idx[0] % len(_IMPORT_TAG_COLORS)]
            color_idx[0] += 1
        cache_key = (tag_l1, tag_l2)
        if cache_key in tag_cache:
            return tag_cache[cache_key]
        parent_id = None
        if tag_l1:
            r1 = await db.execute(select(TagModel).where(TagModel.name == tag_l1))
            l1 = r1.scalar_one_or_none()
            if not l1:
                l1 = TagModel(name=tag_l1, color=color, level=1, parent_id=None)
                db.add(l1)
                await db.flush()
                await db.refresh(l1)
            parent_id = l1.id
        r2 = await db.execute(
            select(TagModel).where(TagModel.name == tag_l2, TagModel.parent_id == parent_id)
        )
        l2 = r2.scalar_one_or_none()
        if not l2:
            l2 = TagModel(name=tag_l2, color=color, level=2, parent_id=parent_id)
            db.add(l2)
            await db.flush()
            await db.refresh(l2)
        tag_cache[cache_key] = l2.id
        return l2.id

    created = 0
    skipped = 0
    invalid = 0
    errors: List[Dict[str, Any]] = []
    existing_hosts = set()
    result = await db.execute(select(ArrayModel.host))
    existing_hosts = {r[0] for r in result.all()}

    for i, row in enumerate(rows):
        norm, reason = _normalize_import_row(row)
        if not norm:
            invalid += 1
            errors.append({"row": i + 2, "reason": reason or "parse_error"})
            logger.info("Import row %s invalid: %s | raw=%s", i + 2, reason, row)
            continue
        host = norm["host"]
        if host in existing_hosts:
            skipped += 1
            errors.append({"row": i + 2, "host": host, "reason": "host_already_exists"})
            continue
        array_id = f"arr_{uuid.uuid4().hex[:8]}"
        try:
            tag_id = await get_or_create_tag_from_import(norm)
            db_array = ArrayModel(
                array_id=array_id,
                name=norm["name"],
                host=host,
                port=norm.get("port", 22),
                username=norm.get("username", "root"),
                saved_password=norm.get("password", ""),
                key_path="",
                folder="",
                tag_id=tag_id,
            )
            db.add(db_array)
            ssh_pool.add_connection(
                array_id=array_id,
                host=host,
                port=norm.get("port", 22),
                username=norm.get("username", "root"),
                password=norm.get("password", ""),
                key_path=None,
            )
            _array_status_cache[array_id] = ArrayStatus(array_id=array_id, name=norm["name"], host=host)
            existing_hosts.add(host)
            created += 1
        except Exception as e:
            logger.exception("Import row %s failed: %s", i + 2, e)
            errors.append({"row": i + 2, "host": host, "reason": str(e)})
    await db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "invalid": invalid,
        "errors": errors,
        "total_rows": len(rows),
    }


@router.post("", response_model=ArrayResponse, status_code=status.HTTP_201_CREATED)
async def create_array(
    array: ArrayCreate,
    db: AsyncSession = Depends(get_db),
    ssh_pool: SSHPool = Depends(get_ssh_pool),
):
    """Create a new array"""
    array_id = f"arr_{uuid.uuid4().hex[:8]}"

    result = await db.execute(select(ArrayModel).where(ArrayModel.host == array.host))
    if result.scalar():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Array with host {array.host} already exists",
        )

    db_array = ArrayModel(
        array_id=array_id,
        name=array.name,
        host=array.host,
        port=array.port,
        username=array.username,
        key_path=array.key_path,
        folder=array.folder,
        tag_id=array.tag_id,
    )
    db.add(db_array)
    await db.commit()
    await db.refresh(db_array)

    ssh_pool.add_connection(
        array_id=array_id,
        host=array.host,
        port=array.port,
        username=array.username,
        password=array.password,
        key_path=array.key_path or None,
    )

    _array_status_cache[array_id] = ArrayStatus(array_id=array_id, name=array.name, host=array.host)
    logger.info(f"Created array: {array.name} ({array.host})")
    return db_array


@router.get("/{array_id}", response_model=ArrayResponse)
async def get_array(
    array_id: str,
    db: AsyncSession = Depends(get_db),
):
    """Get array by ID"""
    result = await db.execute(select(ArrayModel).where(ArrayModel.array_id == array_id))
    array = result.scalar()
    if not array:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Array {array_id} not found",
        )
    return array


@router.put("/{array_id}", response_model=ArrayResponse)
async def update_array(
    array_id: str,
    update: ArrayUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Update array with optimistic locking support."""
    result = await db.execute(select(ArrayModel).where(ArrayModel.array_id == array_id))
    array = result.scalar()
    if not array:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Array {array_id} not found",
        )

    if update.expected_version is not None:
        current_version = getattr(array, 'version', 1) or 1
        if current_version != update.expected_version:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"数据已被其他用户修改，请刷新后重试 (expected version {update.expected_version}, current {current_version})",
            )

    update_data = update.model_dump(exclude_unset=True)
    update_data.pop('expected_version', None)
    for field, value in update_data.items():
        setattr(array, field, value)
    array.version = (getattr(array, 'version', 1) or 1) + 1

    await db.commit()
    await db.refresh(array)

    if array_id in _array_status_cache:
        if update.name:
            _array_status_cache[array_id].name = update.name
        if update.host:
            _array_status_cache[array_id].host = update.host

    return array


@router.delete("/{array_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_array(
    array_id: str,
    db: AsyncSession = Depends(get_db),
    ssh_pool: SSHPool = Depends(get_ssh_pool),
):
    """Delete array"""
    result = await db.execute(select(ArrayModel).where(ArrayModel.array_id == array_id))
    array = result.scalar()
    if not array:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Array {array_id} not found",
        )

    ssh_pool.remove_connection(array_id)
    if array_id in _array_status_cache:
        del _array_status_cache[array_id]
    await db.delete(array)
    await db.commit()
    logger.info(f"Deleted array: {array_id}")


# ---------------------------------------------------------------------------
# Batch action
# ---------------------------------------------------------------------------

@router.post("/batch/{action}")
async def batch_action(
    action: str,
    request: BatchActionRequest,
    stream: bool = Query(False, description="Return SSE progress stream"),
    db: AsyncSession = Depends(get_db),
    ssh_pool: SSHPool = Depends(get_ssh_pool),
):
    """Execute batch operations on multiple arrays."""
    valid_actions = ["connect", "disconnect", "refresh", "deploy-agent", "start-agent", "stop-agent", "restart-agent"]
    if action not in valid_actions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid action. Must be one of: {valid_actions}",
        )

    if not request.array_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No arrays specified",
        )

    array_result = await db.execute(
        select(ArrayModel).where(ArrayModel.array_id.in_(request.array_ids))
    )
    array_map: Dict[str, ArrayModel] = {arr.array_id: arr for arr in array_result.scalars().all()}

    async def _update_saved_password(array_id: str, effective_password: str) -> None:
        if not AsyncSessionLocal or not effective_password:
            return
        try:
            async with AsyncSessionLocal() as session:
                row = await session.execute(select(ArrayModel).where(ArrayModel.array_id == array_id))
                array = row.scalar_one_or_none()
                if array and array.saved_password != effective_password:
                    array.saved_password = effective_password
                    await session.commit()
        except Exception:
            logger.debug("Failed to persist saved password for %s", array_id)

    async def execute_single(array_id: str) -> Dict[str, Any]:
        try:
            array = array_map.get(array_id)
            if not array:
                return {"array_id": array_id, "success": False, "error": "Array not found"}

            conn = ssh_pool.get_connection(array_id)
            config = get_config()

            if action == "connect":
                effective_password = (request.password or getattr(array, "saved_password", "") or "").strip()
                if not conn:
                    conn = ssh_pool.add_connection(
                        array_id=array_id,
                        host=array.host,
                        port=array.port,
                        username=array.username,
                        password=effective_password,
                        key_path=array.key_path or None,
                    )
                else:
                    conn.password = effective_password
                success = await _run_blocking(conn.connect, max(15, get_config().ssh.timeout + 5))
                if success:
                    status_obj = _get_array_status(array_id)
                    status_obj.state = conn.state
                    if effective_password and (not array.saved_password or array.saved_password != effective_password):
                        await _update_saved_password(array_id, effective_password)
                    return {"array_id": array_id, "success": True, "message": "Connected"}
                else:
                    return {"array_id": array_id, "success": False, "error": conn.last_error}

            elif action == "disconnect":
                ssh_pool.disconnect(array_id)
                status_obj = _get_array_status(array_id)
                status_obj.state = ConnectionState.DISCONNECTED
                return {"array_id": array_id, "success": True, "message": "Disconnected"}

            elif action == "refresh":
                if not conn or not conn.is_connected():
                    return {"array_id": array_id, "success": False, "error": "Array not connected"}
                deployer = AgentDeployer(conn, config)
                status_obj = _get_array_status(array_id)
                status_obj.agent_deployed = await _run_blocking(deployer.check_deployed, 10)
                status_obj.agent_running = await _run_blocking(deployer.check_running, 10)
                status_obj.last_refresh = datetime.now()
                return {"array_id": array_id, "success": True, "message": "Refreshed"}

            elif action == "deploy-agent":
                if not conn or not conn.is_connected():
                    return {"array_id": array_id, "success": False, "error": "Array not connected"}
                deployer = AgentDeployer(conn, config)
                result = await _run_blocking(deployer.deploy, 120)
                if result.get("ok"):
                    status_obj = _get_array_status(array_id)
                    status_obj.agent_deployed = True
                    return {"array_id": array_id, "success": True, "message": "Agent deployed", "warnings": result.get("warnings", [])}
                else:
                    return {"array_id": array_id, "success": False, "error": result.get("error")}

            elif action == "start-agent":
                if not conn or not conn.is_connected():
                    return {"array_id": array_id, "success": False, "error": "Array not connected"}
                deployer = AgentDeployer(conn, config)
                result = await _run_blocking(deployer.start_agent, 60)
                if result.get("ok"):
                    status_obj = _get_array_status(array_id)
                    status_obj.agent_running = True
                    return {"array_id": array_id, "success": True, "message": "Agent started"}
                else:
                    return {"array_id": array_id, "success": False, "error": result.get("error")}

            elif action == "stop-agent":
                if not conn or not conn.is_connected():
                    return {"array_id": array_id, "success": False, "error": "Array not connected"}
                deployer = AgentDeployer(conn, config)
                result = await _run_blocking(deployer.stop_agent, 30)
                if result.get("ok"):
                    status_obj = _get_array_status(array_id)
                    status_obj.agent_running = False
                    return {"array_id": array_id, "success": True, "message": "Agent stopped"}
                else:
                    return {"array_id": array_id, "success": False, "error": result.get("error")}

            elif action == "restart-agent":
                if not conn or not conn.is_connected():
                    return {"array_id": array_id, "success": False, "error": "Array not connected"}
                deployer = AgentDeployer(conn, config)
                result = await _run_blocking(deployer.restart_agent, 60)
                if result.get("ok"):
                    status_obj = _get_array_status(array_id)
                    status_obj.agent_running = True
                    return {"array_id": array_id, "success": True, "message": "Agent restarted"}
                else:
                    return {"array_id": array_id, "success": False, "error": result.get("error")}

            return {"array_id": array_id, "success": False, "error": "Unknown action"}

        except Exception as e:
            sys_error("batch", f"Batch action {action} failed for {array_id}", {"error": str(e)})
            return {"array_id": array_id, "success": False, "error": str(e)}

    async def _run_batch() -> Dict[str, Any]:
        max_concurrency = 5
        semaphore = asyncio.Semaphore(max_concurrency)

        async def _with_limit(array_id: str) -> Dict[str, Any]:
            async with semaphore:
                return await execute_single(array_id)

        tasks = [_with_limit(array_id) for array_id in request.array_ids]
        results = await asyncio.gather(*tasks)
        success_count = sum(1 for r in results if r.get("success"))
        sys_info("batch", f"Batch {action} completed", {
            "total": len(request.array_ids),
            "success": success_count,
            "failed": len(request.array_ids) - success_count,
        })
        return {
            "action": action,
            "total": len(request.array_ids),
            "success_count": success_count,
            "results": results,
        }

    if not stream:
        return await _run_batch()

    async def _sse_stream():
        total = len(request.array_ids)
        completed = 0
        success_count = 0
        results = []
        max_concurrency = 5
        semaphore = asyncio.Semaphore(max_concurrency)

        async def _with_limit(array_id: str) -> Dict[str, Any]:
            async with semaphore:
                return await execute_single(array_id)

        yield f"data: {json.dumps({'type': 'start', 'action': action, 'total': total}, ensure_ascii=False)}\n\n"
        tasks = {asyncio.create_task(_with_limit(array_id)): array_id for array_id in request.array_ids}
        for task in asyncio.as_completed(tasks):
            res = await task
            results.append(res)
            completed += 1
            if res.get("success"):
                success_count += 1
            event = {
                "type": "progress",
                "action": action,
                "completed": completed,
                "total": total,
                "success_count": success_count,
                "result": res,
            }
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
        summary = {
            "type": "done",
            "action": action,
            "completed": completed,
            "total": total,
            "success_count": success_count,
            "results": results,
        }
        yield f"data: {json.dumps(summary, ensure_ascii=False)}\n\n"

    return StreamingResponse(_sse_stream(), media_type="text/event-stream")


# ---------------------------------------------------------------------------
# Connect / Disconnect
# ---------------------------------------------------------------------------

@router.post("/{array_id}/connect")
async def connect_array(
    array_id: str,
    password: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    ssh_pool: SSHPool = Depends(get_ssh_pool),
):
    """Connect to array."""
    result = await db.execute(select(ArrayModel).where(ArrayModel.array_id == array_id))
    array = result.scalar()
    if not array:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Array {array_id} not found",
        )

    effective_password = password or getattr(array, 'saved_password', '') or ''

    conn = ssh_pool.get_connection(array_id)
    if not conn:
        conn = ssh_pool.add_connection(
            array_id=array_id,
            host=array.host,
            port=array.port,
            username=array.username,
            password=effective_password,
            key_path=array.key_path or None,
        )
    elif effective_password:
        conn.password = effective_password

    success = await _run_blocking(
        conn.connect,
        max(15, get_config().ssh.timeout + 5),
    )

    status_obj = _get_array_status(array_id)
    status_obj.state = conn.state
    status_obj.last_error = conn.last_error

    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Connection failed: {conn.last_error}",
        )

    if effective_password and (not array.saved_password or array.saved_password != effective_password):
        try:
            array.saved_password = effective_password
            await db.commit()
        except Exception:
            pass

    config = get_config()
    deployer = AgentDeployer(conn, config)
    status_obj.agent_deployed = await _run_blocking(deployer.check_deployed, 10)
    agent_state = await _run_blocking(deployer._resolve_running_state, 15)
    status_obj.agent_running = agent_state["running"]
    status_obj.running_source = agent_state["running_source"]
    status_obj.running_confidence = agent_state["running_confidence"]
    status_obj.service_active = agent_state["service_active"]
    status_obj.service_substate = agent_state.get("service_substate", "")
    status_obj.main_pid = agent_state.get("main_pid")
    status_obj.pidfile_present = agent_state["pidfile_present"]
    status_obj.pidfile_pid = agent_state.get("pidfile_pid")
    status_obj.pidfile_stale = agent_state.get("pidfile_stale", False)
    status_obj.matched_process_cmdline = agent_state.get("matched_process_cmdline", "")

    from .websocket import broadcast_status_update
    await broadcast_status_update(array_id, {
        "array_id": array_id,
        "state": "connected",
        "agent_running": status_obj.agent_running,
        "agent_deployed": status_obj.agent_deployed,
        "agent_healthy": status_obj.agent_healthy,
        "transport_connected": True,
        "event": "connect",
    })

    if not status_obj.agent_deployed:
        return {
            "status": "connected",
            "agent_status": "not_deployed",
            "hint": "Agent 未部署，是否立即部署？",
            "agent_deployed": status_obj.agent_deployed,
            "agent_running": status_obj.agent_running,
            "agent_healthy": status_obj.agent_healthy,
            "has_saved_password": True,
        }

    return {
        "status": "connected",
        "agent_deployed": status_obj.agent_deployed,
        "agent_running": status_obj.agent_running,
        "agent_healthy": status_obj.agent_healthy,
        "has_saved_password": True,
    }


@router.post("/{array_id}/disconnect")
async def disconnect_array(
    array_id: str,
    db: AsyncSession = Depends(get_db),
    ssh_pool: SSHPool = Depends(get_ssh_pool),
):
    """Disconnect from array"""
    await _get_array_or_404(array_id, db)
    ssh_pool.disconnect(array_id)

    status_obj = _get_array_status(array_id)
    status_obj.state = ConnectionState.DISCONNECTED
    status_obj.transport_connected = False

    from .websocket import broadcast_status_update
    await broadcast_status_update(array_id, {
        "array_id": array_id,
        "state": "disconnected",
        "transport_connected": False,
        "event": "disconnect",
    })

    return {"status": "disconnected"}


# ---------------------------------------------------------------------------
# Metrics
# ---------------------------------------------------------------------------

@router.get("/{array_id}/metrics")
async def get_array_metrics(
    array_id: str,
    minutes: int = Query(60, description="Time range in minutes"),
    ssh_pool: SSHPool = Depends(get_ssh_pool),
):
    """Get performance metrics (CPU, memory) from the remote array."""
    from .ingest import get_metrics_for_ip

    conn = ssh_pool.get_connection(array_id)
    if not conn or not conn.is_connected():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Array not connected",
        )

    config = get_config()
    metrics_path = config.remote.agent_log_path.replace('alerts.log', 'metrics.jsonl')
    lines_needed = min(minutes * 6, 2000)
    metrics = []

    exit_code, content, _ = await _run_blocking(
        conn.execute, 12,
        f"tail -n {lines_needed} {metrics_path} 2>/dev/null",
        timeout=10,
    )

    if exit_code == 0 and content and content.strip():
        for line in content.strip().split('\n'):
            if not line.strip():
                continue
            try:
                record = json.loads(line)
                metrics.append(record)
            except Exception:
                pass

    for aid, conn_obj in ssh_pool._connections.items():
        if aid == array_id:
            pushed = get_metrics_for_ip(conn_obj.host, minutes)
            if pushed:
                metrics.extend(pushed)
            break

    metrics.sort(key=lambda m: m.get('ts', ''))

    from datetime import timedelta
    cutoff = (datetime.now() - timedelta(minutes=minutes)).isoformat()
    metrics = [m for m in metrics if m.get('ts', '') >= cutoff]

    return {
        "array_id": array_id,
        "minutes": minutes,
        "count": len(metrics),
        "metrics": metrics,
    }


# ---------------------------------------------------------------------------
# Backward-compat re-exports
# These let existing importers (main.py, alert_sync.py, acknowledgements.py,
# system_alerts.py) keep their current "from .arrays import ..." statements.
# ---------------------------------------------------------------------------

# from .array_status — already imported above
# from .array_alert_sync — already imported above
# _derive_active_issues_from_db is not directly re-exported here because
# core/alert_sync.py imports it from arrays (see below)

from .array_status import (  # noqa: E402, F401
    _derive_active_issues_from_db,
    _get_array_status as _get_array_status,  # already imported above
    _update_active_issues,
    _recovery_timestamps,
    _compute_recent_alert_summary,
)
